#!/usr/bin/env python3
"""Parse Infrastructure Scorecard Excel to generate tier-thresholds.json.

Reads the official Infrastructure Scorecard spreadsheet and outputs a structured
JSON file with tier thresholds for each check. This JSON is used by build-scorecard.py
to determine T1/T2/T3 compliance levels.

Usage:
    python3 scripts/parse-tier-thresholds.py
    python3 scripts/parse-tier-thresholds.py --excel path/to/scorecard.xlsx
"""
import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Default paths
if os.environ.get("GITHUB_WORKSPACE"):
    BASE_DIR = Path(os.environ["GITHUB_WORKSPACE"])
else:
    BASE_DIR = Path(__file__).parent.parent

DEFAULT_EXCEL = BASE_DIR / "data" / "Infrastructure Scorecard.xlsx"
OUTPUT_FILE = BASE_DIR / "scorecard" / "tier-thresholds.json"


def slugify(name: str) -> str:
    """Convert check name to a slug for JSON keys."""
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9]+', '_', slug)
    slug = slug.strip('_')
    return slug


def parse_scope(scope_str: str) -> list:
    """Parse scope string into list."""
    if not scope_str:
        return []
    return [s.strip() for s in scope_str.split(',')]


def parse_excel(excel_path: Path) -> dict:
    """Parse the Infrastructure Scorecard Excel file."""
    print(f"Reading: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)

    if "Checks" not in wb.sheetnames:
        raise ValueError(f"Excel file missing 'Checks' sheet. Found: {wb.sheetnames}")

    ws = wb["Checks"]
    rows = list(ws.iter_rows(values_only=True))

    # Expected headers
    headers = rows[0]
    print(f"Headers: {[h for h in headers if h]}")

    # Column indices (0-based)
    COL_PRIORITY = 0      # P
    COL_SCOPE = 1         # Scope
    COL_SOR = 2           # SOR (Source of Record)
    COL_THEME = 3         # Theme (Pillar)
    COL_CHECK = 4         # Check name
    COL_DATA_PROVIDER = 5 # Data Provider
    COL_TIER1 = 6         # Tier 1 Threshold
    COL_TIER2 = 7         # Tier 2 Threshold
    COL_TIER3 = 8         # Tier 3 Threshold
    COL_DESCRIPTION = 11  # Description

    checks = {}
    pillar_order = []

    for row in rows[1:]:
        # Skip empty rows
        if not row[COL_PRIORITY]:
            continue

        check_name = row[COL_CHECK]
        if not check_name:
            continue

        slug = slugify(check_name)
        theme = row[COL_THEME] or "Unknown"

        # Track pillar order
        if theme not in pillar_order:
            pillar_order.append(theme)

        checks[slug] = {
            "name": check_name,
            "pillar": theme,
            "priority": row[COL_PRIORITY] or "P3",
            "scope": parse_scope(row[COL_SCOPE] or ""),
            "sor": row[COL_SOR] or "",
            "data_provider": row[COL_DATA_PROVIDER] or "",
            "tier1": {
                "threshold": row[COL_TIER1] or "",
            },
            "tier2": {
                "threshold": row[COL_TIER2] or "",
            },
            "tier3": {
                "threshold": row[COL_TIER3] or "",
            },
            "description": row[COL_DESCRIPTION] or "",
        }

    return {
        "version": "Q2-2026",
        "source": excel_path.name,
        "generated": datetime.utcnow().isoformat() + "Z",
        "checks": checks,
        "pillar_order": pillar_order,
        "priority_order": ["P1", "P2", "P3"],
        "status_values": {
            "tier1_met": "T1",
            "tier2_met": "T2",
            "tier3_met": "T3",
            "below_tier3": "<T3",
            "not_applicable": "N/A",
            "unknown": "x"
        }
    }


def main():
    parser = argparse.ArgumentParser(description="Parse Infrastructure Scorecard Excel")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL,
                        help=f"Path to Excel file (default: {DEFAULT_EXCEL})")
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE,
                        help=f"Output JSON path (default: {OUTPUT_FILE})")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"Error: Excel file not found: {args.excel}")
        print("Please download 'Infrastructure Scorecard.xlsx' and place it in data/")
        exit(1)

    # Parse and generate
    data = parse_excel(args.excel)

    # Ensure output directory exists
    args.output.parent.mkdir(parents=True, exist_ok=True)

    # Write JSON
    with open(args.output, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"\nGenerated: {args.output}")
    print(f"  - {len(data['checks'])} checks")
    print(f"  - Pillars: {data['pillar_order']}")

    # Summary by priority
    by_priority = {}
    for check in data['checks'].values():
        p = check['priority']
        by_priority[p] = by_priority.get(p, 0) + 1
    print(f"  - By priority: {by_priority}")


if __name__ == "__main__":
    main()
