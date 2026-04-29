#!/usr/bin/env python3
"""Parse Billing team's Infrastructure Hardening Tracker Excel file.

Extracts summary stats and outputs JSON for the scorecard.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

# Default input file location
DEFAULT_INPUT = Path.home() / "Downloads/[Billing] Infra Hardening Tracker V2.xlsx"
OUTPUT_DIR = Path.home() / "Desktop/Github/team-management/scorecard"


def parse_tier(tier_value):
    """Normalize tier value to standard format."""
    if not tier_value:
        return "unknown"

    tier_str = str(tier_value).strip().lower()

    if "tier 1" in tier_str and "met" in tier_str:
        return "t1"
    elif "tier 2" in tier_str and "met" in tier_str:
        return "t2"
    elif "tier 3" in tier_str and "met" in tier_str:
        return "t3"
    elif "< tier 3" in tier_str or "<tier 3" in tier_str or "below" in tier_str:
        return "below_t3"
    elif "not applicable" in tier_str or "n/a" in tier_str:
        return "n/a"
    else:
        return "unknown"


def parse_billing_tracker(input_file):
    """Parse the Billing tracker Excel file and return summary stats."""
    wb = load_workbook(input_file)
    ws = wb.active

    services = {}  # service_name -> {checks: {check_name: tier}}

    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        service_name = row[0]
        check_name = row[3]
        tier_value = row[4]

        if not service_name or not check_name:
            continue

        if service_name not in services:
            services[service_name] = {
                "name": service_name,
                "lang": row[1],
                "repo": row[2],
                "checks": {}
            }

        tier = parse_tier(tier_value)
        services[service_name]["checks"][check_name] = {
            "tier": tier,
            "status": row[5],
            "jira": row[7]
        }

    # Calculate summary stats
    total_checks = 0
    tier_counts = {"t1": 0, "t2": 0, "t3": 0, "below_t3": 0, "n/a": 0, "unknown": 0}

    for service in services.values():
        for check_name, check_data in service["checks"].items():
            tier = check_data["tier"]
            tier_counts[tier] += 1
            total_checks += 1

    # Calculate P1 compliance (T1 / (total - n/a - unknown))
    scorable_checks = total_checks - tier_counts["n/a"] - tier_counts["unknown"]
    compliance_pct = (tier_counts["t1"] / scorable_checks * 100) if scorable_checks > 0 else 0

    return {
        "services": list(services.values()),
        "service_count": len(services),
        "summary": {
            "total_checks": total_checks,
            "t1": tier_counts["t1"],
            "t2": tier_counts["t2"],
            "t3": tier_counts["t3"],
            "below_t3": tier_counts["below_t3"],
            "n_a": tier_counts["n/a"],
            "unknown": tier_counts["unknown"],
            "scorable_checks": scorable_checks,
            "compliance_pct": round(compliance_pct, 1)
        }
    }


def main():
    input_file = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT

    if not input_file.exists():
        print(f"Error: File not found: {input_file}", file=sys.stderr)
        print(f"Usage: {sys.argv[0]} [path/to/billing-tracker.xlsx]", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {input_file}", file=sys.stderr)

    result = parse_billing_tracker(input_file)

    # Save output
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / "billing-tracker.json"
    with open(output_file, "w") as f:
        json.dump(result, f, indent=2)

    # Print summary
    summary = result["summary"]
    print(f"\n{'='*50}")
    print(f"  BILLING TEAM INFRASTRUCTURE HARDENING")
    print(f"{'='*50}")
    print(f"  Services: {result['service_count']}")
    print(f"  Total Checks: {summary['total_checks']}")
    print(f"{'='*50}")
    print(f"  T1 (Tier 1 Met):    {summary['t1']}")
    print(f"  T2 (Tier 2 Met):    {summary['t2']}")
    print(f"  T3 (Tier 3 Met):    {summary['t3']}")
    print(f"  <T3 (Below Tier 3): {summary['below_t3']}")
    print(f"  N/A:                {summary['n_a']}")
    print(f"  Unknown:            {summary['unknown']}")
    print(f"{'='*50}")
    print(f"  P1 Tier 1 Compliance: {summary['compliance_pct']}% ({summary['t1']}/{summary['scorable_checks']})")
    print(f"{'='*50}")
    print(f"\nWrote: {output_file}")


if __name__ == "__main__":
    main()
